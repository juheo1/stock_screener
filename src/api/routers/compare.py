"""
src.api.routers.compare
=======================
POST /compare             -- side-by-side metric comparison for up to 50 tickers.
POST /compare/export.xlsx -- Excel export.
"""

from __future__ import annotations

import io
import logging
from datetime import date

import yfinance as yf
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from src.api.deps import get_db
from src.api.schemas import CompareCell, CompareRequest, CompareResponse, CompareRow
from src.metrics import get_latest_metrics
from src.models import Equity, Flag

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/compare", tags=["compare"])

# Colour band thresholds (percentile-based heuristics)
_BANDS = {
    "gross_margin":       {"good": 40.0, "bad": 15.0},   # %
    "roic":               {"good": 0.12, "bad": 0.05},   # ratio
    "fcf_margin":         {"good": 10.0, "bad": 0.0},    # %
    "interest_coverage":  {"good": 3.0,  "bad": 1.0},   # ratio
    "pe_ratio":           {"good": 15.0, "bad": 40.0},   # lower is better
}


def _band_for(metric: str, value: float | None) -> str:
    """Return a colour band string for a metric value.

    Parameters
    ----------
    metric : str
        Metric name key in :data:`_BANDS`.
    value : float | None

    Returns
    -------
    str
        ``"good"``, ``"bad"``, or ``"neutral"``; ``"na"`` if value is None.
    """
    if value is None:
        return "na"
    thresholds = _BANDS.get(metric)
    if thresholds is None:
        return "neutral"

    good = thresholds["good"]
    bad = thresholds["bad"]

    if metric == "pe_ratio":
        # Lower P/E is better
        if value <= good:
            return "good"
        if value >= bad:
            return "bad"
    else:
        if value >= good:
            return "good"
        if value <= bad:
            return "bad"
    return "neutral"


def _overall_score(row: CompareRow) -> float:
    """Compute a 0–100 composite score from pass/fail bands.

    Parameters
    ----------
    row : CompareRow

    Returns
    -------
    float
    """
    cells = [row.gross_margin, row.roic, row.fcf_margin, row.interest_coverage, row.pe_ratio]
    score = 0.0
    for cell in cells:
        if cell.band == "good":
            score += 20
        elif cell.band == "neutral":
            score += 10
    return score


@router.post("", response_model=CompareResponse)
def compare_tickers(
    request: CompareRequest,
    db: Session = Depends(get_db),
) -> CompareResponse:
    """Return a side-by-side metric comparison for up to 50 tickers.

    Parameters
    ----------
    request : CompareRequest
        ``tickers`` list plus optional threshold overrides.
    db : Session
        Injected database session.

    Returns
    -------
    CompareResponse
    """
    rows: list[CompareRow] = []
    thresholds = {
        "min_gross_margin": request.min_gross_margin,
        "min_roic": request.min_roic,
        "min_fcf_margin": request.min_fcf_margin,
        "min_interest_coverage": request.min_interest_coverage,
        "max_pe": request.max_pe,
    }

    for ticker_sym in request.tickers[:50]:
        ticker_sym = ticker_sym.upper().strip()
        m = get_latest_metrics(ticker_sym, db)
        eq = db.get(Equity, ticker_sym)
        latest_flag = (
            db.query(Flag)
            .filter_by(ticker=ticker_sym, is_zombie=True)
            .order_by(Flag.asof_date.desc())
            .first()
        )

        gm_val = m.gross_margin if m else None
        roic_val = m.roic if m else None
        fcf_val = m.fcf_margin if m else None
        ic_val = m.interest_coverage if m else None
        pe_val = m.pe_ratio if m else None

        compare_row = CompareRow(
            ticker=ticker_sym,
            name=eq.name if eq else ticker_sym,
            sector=eq.sector if eq else None,
            gross_margin=CompareCell(value=gm_val, band=_band_for("gross_margin", gm_val)),
            roic=CompareCell(value=roic_val, band=_band_for("roic", roic_val)),
            fcf_margin=CompareCell(value=fcf_val, band=_band_for("fcf_margin", fcf_val)),
            interest_coverage=CompareCell(value=ic_val, band=_band_for("interest_coverage", ic_val)),
            pe_ratio=CompareCell(value=pe_val, band=_band_for("pe_ratio", pe_val)),
            is_zombie=latest_flag is not None,
        )
        compare_row.overall_score = _overall_score(compare_row)
        rows.append(compare_row)

    return CompareResponse(rows=rows, thresholds=thresholds)


@router.post("/export.xlsx")
def export_compare_xlsx(
    request: CompareRequest,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export compare results as an Excel workbook.

    Parameters
    ----------
    request : CompareRequest
    db : Session

    Returns
    -------
    StreamingResponse
        ``application/vnd.openxmlformats-officedocument.spreadsheetml.sheet``
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    result = compare_tickers(request, db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compare"

    headers = ["Ticker", "Name", "Sector", "Gross Margin %", "ROIC",
               "FCF Margin %", "Interest Coverage", "P/E Ratio", "Overall Score", "Zombie"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    FILL_GOOD = PatternFill("solid", fgColor="1A7A4A")
    FILL_BAD = PatternFill("solid", fgColor="B22222")
    FILL_NA = PatternFill("solid", fgColor="555555")

    band_fills = {"good": FILL_GOOD, "bad": FILL_BAD, "na": FILL_NA}

    metric_cols = {
        "gross_margin": 4,
        "roic": 5,
        "fcf_margin": 6,
        "interest_coverage": 7,
        "pe_ratio": 8,
    }

    for row in result.rows:
        ws.append([
            row.ticker,
            row.name,
            row.sector,
            row.gross_margin.value,
            row.roic.value,
            row.fcf_margin.value,
            row.interest_coverage.value,
            row.pe_ratio.value,
            row.overall_score,
            "YES" if row.is_zombie else "no",
        ])
        excel_row = ws.max_row
        for metric_name, col_idx in metric_cols.items():
            cell_val = getattr(row, metric_name)
            fill = band_fills.get(cell_val.band)
            if fill:
                ws.cell(row=excel_row, column=col_idx).fill = fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=compare_results.xlsx"},
    )
